from openpyxl import load_workbook

def runLog(EXCEL_NAME , datas):
    excel = load_workbook(EXCEL_NAME)
    excelSheet = excel.active
    max_row = excelSheet.max_row+1
    data_row =1
    for datalist in datas:
        excelSheet.cell(row = max_row+(data_row-1) ,column=1,value=datalist[0])
        excelSheet.cell(row = max_row+(data_row-1) ,column=2,value=datalist[1])
        excelSheet.cell(row = max_row+(data_row-1) ,column=3,value=datalist[2])
        excelSheet.cell(row = max_row+(data_row-1) ,column=4,value=datalist[3])
        excelSheet.cell(row = max_row+(data_row-1) ,column=5,value=datalist[4])
        data_row+=1
    

    excel.save(EXCEL_NAME)
    return

def runfile(EXCEL_NAME,files):
    excel = load_workbook(EXCEL_NAME)
    excelSheet = excel["ReadFileList"]
    max_row = excelSheet.max_row+1
    data_row =1
    
    for _file in files:
        excelSheet.cell(row = max_row+(data_row-1) ,column=1,value=_file)
        data_row+=1
    excel.save(EXCEL_NAME)
    return