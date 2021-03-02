from openpyxl import load_workbook
from openpyxl import Workbook
import os
from openpyxl.styles import PatternFill, Color





#현제 경로의 파일을 모두 읽어서 엑셀이 있다면 열고 없다면 새로 생성한다.
def openExcel(EXCEL_NAME):
    fileList= os.listdir()
    for target in fileList:
        if(target == EXCEL_NAME):
            return 


    sheetPatternFill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')

    excel = Workbook()
    excelSheet = excel.active
    excelSheet.freeze_panes = "A2" #행고정
    excelSheet.paper_size
    excelSheet.title ="BlackLogs"
    excelSheet["A1"]="Time"
    excelSheet["A1"].fill = sheetPatternFill
    excelSheet.column_dimensions["A"].width =30

    excelSheet["B1"]="ip"
    excelSheet["B1"].fill = sheetPatternFill
    excelSheet.column_dimensions["B"].width =20

    excelSheet["C1"]="url"
    excelSheet["C1"].fill = sheetPatternFill
    excelSheet.column_dimensions["C"].width =50

    excelSheet["D1"]="query"
    excelSheet["D1"].fill = sheetPatternFill
    excelSheet.column_dimensions["D"].width =50

    excelSheet["E1"]="body"
    excelSheet["E1"].fill = sheetPatternFill
    excelSheet.column_dimensions["E"].width =50





    excel.create_sheet("AttackCount") # 시트 생성과 동시에 이름 지정
    excelSheet2 = excel["AttackCount"]
    excelSheet2["A1"]="Attack_IP"
    excelSheet2.freeze_panes = "A2" #행고정
    excelSheet2["A1"].fill = sheetPatternFill
    excelSheet2.column_dimensions["A"].width =30

    excelSheet2["B1"]="Count"
    excelSheet2["B1"].fill = sheetPatternFill
    excelSheet2.column_dimensions["B"].width =20
    


    excel.create_sheet("ReadFileList") # 시트 생성과 동시에 이름 지정
    excelSheet3 = excel["ReadFileList"]
    excelSheet3["A1"]="FileName"
    excelSheet3.freeze_panes = "A2" #행고정
    excelSheet3["A1"].fill = sheetPatternFill
    excelSheet3.column_dimensions["A"].width =30
    excel.save(EXCEL_NAME)

    



def run(EXCEL_NAME):
    openExcel(EXCEL_NAME)
    
