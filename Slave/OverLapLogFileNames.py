from openpyxl import load_workbook
    

def run(EXCEL_NAME):
    result = []
    excel = load_workbook(EXCEL_NAME)
    excelSheet = excel["ReadFileList"]
    max_row = excelSheet.max_row
    for index in range(2,max_row+1):
        result.append(excelSheet.cell(row = index ,column=1).value)
    
    return result