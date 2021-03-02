
from openpyxl import load_workbook


# 엑셀을 이용하여 데이터를 얻는다.

def attackCount(EXCEL_NAME):
    ipList = getAttackIpList(EXCEL_NAME)
    attackDic = ipCounting(ipList)
    saveExcel(EXCEL_NAME,attackDic)


def getAttackIpList(EXCEL_NAME):
    result = []
    excel = load_workbook(EXCEL_NAME)
    excelSheet = excel["BlackLogs"]
    max_row = excelSheet.max_row
    for index in range(2,max_row+1):
        result.append(excelSheet.cell(row = index ,column=2).value)
    
    return result


def ipCounting(iplist):
    attackDic ={}
    for ip in iplist:
        attackDic[ip]= (attackDic.get(ip,0)+1)
    return attackDic


def saveExcel(EXCEL_NAME,attackDic):
    excel = load_workbook(EXCEL_NAME)
    excelSheet = excel["AttackCount"]
    index =2
    for dic_key,dic_value in attackDic.items():
        excelSheet.cell(row = index ,column=1,value=dic_key)
        excelSheet.cell(row = index ,column=2,value=dic_value)
        index+=1

    excel.save(EXCEL_NAME)
